import subprocess
import json
import openpyxl
import io

GIT = r"C:\Users\USUARIO\AppData\Local\GitHubDesktop\app-3.6.3\resources\app\git\cmd\git.exe"

def run_git(args):
    res = subprocess.run([GIT] + args, capture_output=True)
    return res.stdout

print("=== CHECKING ALL GIT COMMITS OF Base de datos Admin.xlsx FOR COL J (INQUILINO) ===")
res_commits = run_git(["log", "--oneline", "--", "Base de datos Admin.xlsx"])
commits = [l.decode('utf-8', errors='ignore').split()[0] for l in res_commits.strip().splitlines() if l]

print(f"Total commits of Base de datos Admin.xlsx: {len(commits)}")

found_excel_tenants = {}

for c in commits:
    excel_bytes = run_git(["show", f"{c}:Base de datos Admin.xlsx"])
    if excel_bytes and len(excel_bytes) > 1000:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            for sname in wb.sheetnames:
                if 'ADMIN' in sname.upper():
                    ws = wb[sname]
                    for r in range(6, ws.max_row + 1):
                        prop_name = ws.cell(r, 9).value
                        tenant_val = ws.cell(r, 10).value
                        if prop_name and tenant_val and str(tenant_val).strip() and str(tenant_val).strip().lower() != 'none':
                            p_str = str(prop_name).strip()
                            t_str = str(tenant_val).strip()
                            if p_str not in found_excel_tenants:
                                found_excel_tenants[p_str] = set()
                            found_excel_tenants[p_str].add((c, t_str))
        except Exception as e:
            pass

print(f"Found tenant names in Excel git history for {len(found_excel_tenants)} properties:")
for p, tset in found_excel_tenants.items():
    print(f"  Property '{p}':")
    for commit, tname in tset:
        print(f"    -> Commit {commit}: '{tname}'")
