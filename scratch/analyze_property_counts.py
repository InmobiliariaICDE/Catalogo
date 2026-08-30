import json
import openpyxl
import requests

print("=== 1. LOCAL EXCEL PROPERTIES (Base de datos Admin.xlsx) ===")
wb = openpyxl.load_workbook('Base de datos Admin.xlsx', data_only=True)
ws = wb['ADMINISTRACION DETALLADA']

excel_props = []
for r in range(6, ws.max_row + 1):
    pid = ws.cell(r, 1).value
    pname = ws.cell(r, 9).value
    powner = ws.cell(r, 7).value
    ptenant = ws.cell(r, 10).value
    prent = ws.cell(r, 17).value
    if pid is not None and pname is not None:
        excel_props.append({
            'row': r,
            'id': str(pid).strip(),
            'name': str(pname).strip(),
            'owner': str(powner).strip() if powner else '',
            'tenant': str(ptenant).strip() if ptenant else '',
            'rent': prent
        })

print(f"Excel has {len(excel_props)} properties:")
for p in excel_props:
    print(f"  ID {p['id']:<4} | Row {p['row']:<2} | {p['name']:<40} | Owner: {p['owner']:<20} | Rent: {p['rent']}")

print("\n=== 2. GOOGLE APPS SCRIPT CLOUD PROPERTIES ===")
url = "https://script.google.com/macros/s/AKfycbwAUUSYRhDX6Eik4KA-B6luk74YjCNRanwv13CmmZg4La8NzVuNyBC0T5GH6f4-ke-Xig/exec?action=getAdminData"
try:
    res = requests.get(url, timeout=10)
    cloud_data = res.json()
    cloud_props = cloud_data.get('properties', [])
    cloud_ids = {str(p.get('id')).strip() for p in cloud_props}
    print(f"Cloud has {len(cloud_props)} properties:")
    for p in cloud_props:
        print(f"  ID {str(p.get('id')):<4} | {p.get('name'):<40} | Owner: {p.get('owner'):<20} | Rent: {p.get('monthly_rent')}")
except Exception as e:
    print("Error fetching cloud data:", e)
    cloud_ids = set()

excel_ids = {p['id'] for p in excel_props}
print(f"\nProperties in Excel but MISSING in Cloud: {excel_ids - cloud_ids}")
for p in excel_props:
    if p['id'] in (excel_ids - cloud_ids):
        print(f"  -> MISSING IN CLOUD: ID {p['id']} - {p['name']} (Owner: {p['owner']})")

print(f"\nProperties in Cloud but MISSING in Excel: {cloud_ids - excel_ids}")
for p in cloud_props:
    cid = str(p.get('id')).strip()
    if cid in (cloud_ids - excel_ids):
        print(f"  -> IN CLOUD ONLY: ID {cid} - {p.get('name')} (Owner: {p.get('owner')})")

print("\n=== 3. CATALOG MATCHES FOR PORTAL DEL CAMPO & NOGALES ===")
with open('datos_catalogo.json', 'r', encoding='utf-8') as f:
    cat = json.load(f)

for p in cat:
    code = str(p.get('Código') or p.get('codigo'))
    name = str(p.get('Nombre') or p.get('titulo'))
    if code in ['1338', '325', '463']:
        print(f"  Code {code:<5} | {name:<55} | Adm: {p.get('Administración')} | Price: {p.get('Precio')}")
