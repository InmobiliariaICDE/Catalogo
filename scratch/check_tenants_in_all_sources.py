import openpyxl
import json
import os
import glob

print("=== CHECKING ALL EXCEL FILES FOR TENANT NAMES IN COL J OR ANY COL ===")

for fx in glob.glob("*.xlsx"):
    try:
        wb = openpyxl.load_workbook(fx, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for r in range(1, ws.max_row + 1):
                col_j = ws.cell(r, 10).value
                if col_j and str(col_j).strip() and str(col_j).strip() not in ['Inquilino', 'None', '-']:
                    print(f"File '{fx}' Sheet '{sheet}' Row {r} Col J: '{col_j}' (Prop in Col I: '{ws.cell(r, 9).value}')")
    except Exception as e:
        print(f"Error reading {fx}:", e)
