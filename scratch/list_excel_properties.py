import openpyxl

wb = openpyxl.load_workbook('Base de datos Admin.xlsx', data_only=True)
ws = wb['ADMINISTRACION DETALLADA']

print(f"Total rows in sheet: {ws.max_row}")
props = []

for r in range(1, ws.max_row + 1):
    row_id = ws.cell(r, 1).value
    col7_val = ws.cell(r, 7).value # Owner or code
    col9_val = ws.cell(r, 9).value # Property name
    col10_val = ws.cell(r, 10).value # Tenant name
    col17_val = ws.cell(r, 17).value # Rent
    
    if any(v is not None for v in [row_id, col7_val, col9_val, col10_val]):
        props.append((r, row_id, col7_val, col9_val, col10_val, col17_val))

print(f"Non-empty rows in Excel: {len(props)}")
for p in props:
    print(f"Row {p[0]:2d} | ID: {str(p[1]):<5} | Owner/Col7: {str(p[2]):<20} | Name/Col9: {str(p[3]):<35} | Tenant/Col10: {str(p[4]):<20} | Rent: {str(p[5])}")
