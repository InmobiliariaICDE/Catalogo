import openpyxl
import json
import os
import glob

print("=== 1. SEARCHING ALL XLSX FILES FOR TENANT NAMES IN ALL SHEETS ===")
xlsx_files = glob.glob("*.xlsx")

for fx in xlsx_files:
    print(f"\n--- FILE: {fx} ---")
    try:
        wb = openpyxl.load_workbook(fx, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows_found = 0
            for r in range(1, ws.max_row + 1):
                row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None]
                if row_vals:
                    row_str = " | ".join([str(v) for v in row_vals])
                    if any(k in row_str.lower() for k in ['inquilino', 'arrendatario', 'arriendo', 'apto', 'limonar', 'goya', 'casa azul', 'silvia']):
                        rows_found += 1
                        if rows_found <= 15:
                            print(f"  [{sname} Row {r}]: {row_str[:160]}")
            if rows_found > 15:
                print(f"  ... and {rows_found - 15} more rows in sheet '{sname}'")
    except Exception as e:
        print(f"Error reading {fx}:", e)

print("\n=== 2. SEARCHING JSON & HTML FILES FOR INQUILINO / ARRENDATARIO ===")
for filename in ['leads.json', 'citas.json', 'datos_catalogo.json', 'adminreferenciavieja.html', 'crm_clean.html']:
    if os.path.exists(filename):
        with open(filename, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            for kw in ['inquilino', 'arrendatario', 'tenant']:
                c = content.lower().count(kw)
                if c > 0:
                    print(f"{filename} -> Keyword '{kw}': {c} matches")
