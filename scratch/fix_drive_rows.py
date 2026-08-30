import json
import openpyxl
import requests

# Load admin_data.json
with open('admin_data.json', 'r', encoding='utf-8') as f:
    admin_data = json.load(f)

properties = admin_data.get('properties', [])

# Find or update ID 26 (Los Nogales) and ID 27 (Portal del Campo)
months_list = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
years = ["2023", "2024", "2025", "2026", "2027"]

def make_vacant_payments():
    pay_dict = {}
    for y in years:
        pay_dict[y] = [{"month": m, "value": "DESOCUPADO", "status": "VACANT"} for m in months_list]
    return pay_dict

prop_26 = None
prop_27 = None

for p in properties:
    if str(p.get('id')) == '26' or 'NOGALES' in str(p.get('name')).upper():
        prop_26 = p
    elif str(p.get('id')) == '27' or 'PORTAL DEL CAMPO' in str(p.get('name')).upper():
        prop_27 = p

# Fix prop 26 (APARTAMENTO LOS NOGALES)
if not prop_26:
    prop_26 = {}
    properties.append(prop_26)

prop_26['id'] = '26'
prop_26['excel_row'] = 25
prop_26['owner'] = 'Joel Torres Hurtado'
prop_26['owner_phone'] = '3105641885'
prop_26['name'] = 'APARTAMENTO LOS NOGALES'
prop_26['tenant_name'] = ''
prop_26['tenant_phone'] = ''
prop_26['increase_notes'] = ''
prop_26['damage_notes'] = 'Código Catálogo 463'
prop_26['duration'] = '12'
prop_26['deposit'] = '359000'
prop_26['start_date'] = ''
prop_26['due_day'] = 5.0
prop_26['max_due_day'] = 10.0
prop_26['monthly_rent'] = 359000.0
prop_26['status'] = 'Desocupado'
prop_26['payments'] = make_vacant_payments()

# Fix prop 27 (CASA CONDOMINIO PORTAL DEL CAMPO)
if not prop_27:
    prop_27 = {}
    properties.append(prop_27)

prop_27['id'] = '27'
prop_27['excel_row'] = 26
prop_27['owner'] = 'Inmobiliaria ICDE'
prop_27['owner_phone'] = '3103914892'
prop_27['name'] = 'CASA CONDOMINIO PORTAL DEL CAMPO'
prop_27['tenant_name'] = ''
prop_27['tenant_phone'] = ''
prop_27['increase_notes'] = ''
prop_27['damage_notes'] = 'Código Catálogo 1338 / 325'
prop_27['duration'] = '12'
prop_27['deposit'] = '685000'
prop_27['start_date'] = ''
prop_27['due_day'] = 5.0
prop_27['max_due_day'] = 10.0
prop_27['monthly_rent'] = 685000.0
prop_27['status'] = 'Desocupado'
prop_27['payments'] = make_vacant_payments()

# Sort properties by ID
sorted_props = sorted(properties, key=lambda x: int(x['id']))
admin_data['properties'] = sorted_props

# Save admin_data.json
with open('admin_data.json', 'w', encoding='utf-8') as f:
    json.dump(admin_data, f, indent=4, ensure_ascii=False)

print("Saved admin_data.json with fixed Nogales (Row 26, Owner: Joel Torres Hurtado) and Portal del Campo (Row 27, Owner: Inmobiliaria ICDE).")

# Update Base de datos Admin.xlsx
wb = openpyxl.load_workbook('Base de datos Admin.xlsx')
ws = wb['ADMINISTRACION DETALLADA']

# Row 26 (Index 26 in Excel, ID 26)
r26 = 26
ws.cell(r26, 1, value=26)
ws.cell(r26, 7, value='Joel Torres Hurtado')
ws.cell(r26, 8, value='3105641885')
ws.cell(r26, 9, value='APARTAMENTO LOS NOGALES')
ws.cell(r26, 10, value='')
ws.cell(r26, 17, value=359000)

# Row 27 (Index 27 in Excel, ID 27)
r27 = 27
ws.cell(r27, 1, value=27)
ws.cell(r27, 7, value='Inmobiliaria ICDE')
ws.cell(r27, 8, value='3103914892')
ws.cell(r27, 9, value='CASA CONDOMINIO PORTAL DEL CAMPO')
ws.cell(r27, 10, value='')
ws.cell(r27, 17, value=685000)

wb.save('Base de datos Admin.xlsx')
print("Saved Base de datos Admin.xlsx!")

# Push payload to Google Apps Script cloud to update Drive Sheet
url = "https://script.google.com/macros/s/AKfycbwAUUSYRhDX6Eik4KA-B6luk74YjCNRanwv13CmmZg4La8NzVuNyBC0T5GH6f4-ke-Xig/exec"
payload = {
    "action": "importAdminData",
    "data": {
        "properties": sorted_props
    }
}

try:
    res = requests.post(url, json=payload, timeout=20)
    print("Google Apps Script response:", res.status_code, res.text)
except Exception as e:
    print("Error sending to Google Apps Script:", e)
