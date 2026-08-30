import json
import openpyxl
import requests

# Load admin_data.json
with open('admin_data.json', 'r', encoding='utf-8') as f:
    admin_data = json.load(f)

properties = admin_data.get('properties', [])
months_list = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
years = ["2023", "2024", "2025", "2026", "2027"]

updated_count = 0
for p in properties:
    name_upper = p.get('name', '').upper()
    if 'NOGALES' in name_upper or 'PORTAL DEL CAMPO' in name_upper:
        print(f"Updating property: ID={p.get('id')}, Name='{p.get('name')}' to DESOCUPADO / VACANT")
        p['status'] = 'Desocupado'
        p['tenant_name'] = ''
        p['tenant_phone'] = ''
        p['start_date'] = ''
        
        # Reset payments history for all years to VACANT / DESOCUPADO
        payments_dict = {}
        for y in years:
            payments_dict[y] = [{
                "month": m,
                "value": "DESOCUPADO",
                "status": "VACANT"
            } for m in months_list]
        p['payments'] = payments_dict
        updated_count += 1

print(f"Updated {updated_count} properties in admin_data.json.")

# Save admin_data.json
with open('admin_data.json', 'w', encoding='utf-8') as f:
    json.dump(admin_data, f, indent=4, ensure_ascii=False)

# Update Base de datos Admin.xlsx
wb = openpyxl.load_workbook('Base de datos Admin.xlsx')
ws = wb['ADMINISTRACION DETALLADA']

# Column mappings for years
years_col_map = {
    2023: list(range(18, 30)),
    2024: list(range(31, 43)),
    2025: list(range(44, 56)),
    2026: list(range(57, 69)),
    2027: list(range(70, 82))
}

for r in range(6, ws.max_row + 1):
    prop_name = ws.cell(r, 9).value
    if prop_name and ('NOGALES' in str(prop_name).upper() or 'PORTAL DEL CAMPO' in str(prop_name).upper()):
        ws.cell(r, 10, value='') # Tenant name empty
        ws.cell(r, 11, value='') # Tenant phone empty
        ws.cell(r, 14, value='') # Start date empty
        
        # Set DESOCUPADO for all payment columns
        for y, cols in years_col_map.items():
            for c in cols:
                ws.cell(r, c, value='DESOCUPADO')

wb.save('Base de datos Admin.xlsx')
print("Saved Base de datos Admin.xlsx with DESOCUPADO for Nogales & Portal del Campo.")

# Push to Google Apps Script Cloud via POST importAdminData
url = "https://script.google.com/macros/s/AKfycbwAUUSYRhDX6Eik4KA-B6luk74YjCNRanwv13CmmZg4La8NzVuNyBC0T5GH6f4-ke-Xig/exec"
payload = {
    "action": "importAdminData",
    "data": {
        "properties": properties
    }
}

try:
    res = requests.post(url, json=payload, timeout=20)
    print("Google Apps Script POST response:", res.status_code, res.text)
except Exception as e:
    print("Error posting to Google Apps Script:", e)
