import json
import openpyxl

with open('admin_data.json', 'r', encoding='utf-8') as f:
    admin_data = json.load(f)

properties = admin_data.get('properties', [])

wb = openpyxl.load_workbook('Base de datos Admin.xlsx')
sheet_name = 'ADMINISTRACION DETALLADA'
ws = wb[sheet_name]

# Update or insert rows for all properties
for idx, p in enumerate(properties):
    row_idx = 6 + idx # Rows start at index 6
    ws.cell(row=row_idx, column=1, value=int(p['id'])) # Col A: ID
    ws.cell(row=row_idx, column=2, value=idx + 1)      # Col B: Item
    ws.cell(row=row_idx, column=3, value=int(p['id'])) # Col C: ID
    ws.cell(row=row_idx, column=4, value=idx + 1)      # Col D: Item
    ws.cell(row=row_idx, column=6, value=p.get('damage_notes', '')) # Col F: Daños
    ws.cell(row=row_idx, column=7, value=p.get('owner', ''))        # Col G: Propietario
    ws.cell(row=row_idx, column=8, value=p.get('owner_phone', ''))  # Col H: Celular Propietario
    
    raw_name = p.get('name', '')
    if p.get('increase_notes'):
        raw_name += "  " + p['increase_notes']
    ws.cell(row=row_idx, column=9, value=raw_name)                 # Col I: Inmueble
    
    ws.cell(row=row_idx, column=10, value=p.get('tenant_name', '')) # Col J: Inquilino
    ws.cell(row=row_idx, column=11, value=p.get('tenant_phone', ''))# Col K: Celular Inquilino
    ws.cell(row=row_idx, column=12, value=p.get('duration', '12'))  # Col L: Contrato
    ws.cell(row=row_idx, column=13, value=p.get('deposit', ''))     # Col M: Depósito
    ws.cell(row=row_idx, column=14, value=p.get('start_date', ''))   # Col N: Fecha Inicio
    ws.cell(row=row_idx, column=15, value=p.get('due_day', 5))       # Col O: Día Pago
    ws.cell(row=row_idx, column=16, value=p.get('max_due_day', 10))  # Col P: Límite Pago
    ws.cell(row=row_idx, column=17, value=p.get('monthly_rent', 0))  # Col Q: Canon

wb.save('Base de datos Admin.xlsx')
print(f"SUCCESS: Saved {len(properties)} properties to Base de datos Admin.xlsx!")
