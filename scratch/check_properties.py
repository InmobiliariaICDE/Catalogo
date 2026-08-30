import json
import os
import glob
import openpyxl

print("--- SEARCHING datos_catalogo.json ---")
with open('datos_catalogo.json', 'r', encoding='utf-8') as f:
    cat_data = json.load(f)

print("Total catalog items:", len(cat_data))
for p in cat_data:
    t = str(p.get('titulo', '')).lower()
    c = str(p.get('codigo', '')).lower()
    if 'portal' in t or 'nogal' in t:
        print(f"Code: {p.get('codigo')}, Title: {p.get('titulo')}, Gestion: {p.get('gestion')}, Tipo: {p.get('tipo')}, Estado: {p.get('estado')}")

print("\n--- SEARCHING admin_data.json ---")
with open('admin_data.json', 'r', encoding='utf-8') as f:
    admin_data = json.load(f)

if isinstance(admin_data, list):
    print("Total admin_data items:", len(admin_data))
    for p in admin_data:
        name = p.get('name', '') or p.get('owner', '')
        if 'portal' in name.lower() or 'nogal' in name.lower():
            print("Found in admin_data list:", p)
elif isinstance(admin_data, dict):
    print("admin_data keys:", list(admin_data.keys()))
    for k, v in admin_data.items():
        if isinstance(v, list):
            print(f"Key '{k}' has {len(v)} items")
            for item in v:
                if isinstance(item, dict):
                    name = str(item.get('name', '')) + " " + str(item.get('owner', '')) + " " + str(item.get('tenant_name', ''))
                    if 'portal' in name.lower() or 'nogal' in name.lower():
                        print(f"  Match in {k}: {item.get('id')} - {item.get('name')} (Owner: {item.get('owner')})")

print("\n--- CHECKING Base de datos Admin.xlsx ---")
if os.path.exists('Base de datos Admin.xlsx'):
    wb = openpyxl.load_workbook('Base de datos Admin.xlsx', data_only=True)
    print("Sheets in Base de datos Admin.xlsx:", wb.sheetnames)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        row_count = ws.max_row
        print(f"Sheet '{sheet}': {row_count} rows")
        for r in range(1, row_count + 1):
            row_vals = [str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            row_str = " ".join(row_vals).lower()
            if 'portal' in row_str or 'nogal' in row_str:
                print(f"  Match row {r} in sheet '{sheet}':", row_vals[:5])

print("\n--- CHECKING HTML files (administramos-casas-en-arriendo-neiva.html, admin.html) ---")
for html_file in ['administramos-casas-en-arriendo-neiva.html', 'admin.html', 'index.html', 'categorizador.html']:
    if os.path.exists(html_file):
        with open(html_file, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            count_portal = content.lower().count('portal del campo')
            count_nogales = content.lower().count('nogales')
            print(f"{html_file}: 'portal del campo' count={count_portal}, 'nogales' count={count_nogales}")
