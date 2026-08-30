import json
import openpyxl
import requests

# Load admin_data.json
with open('admin_data.json', 'r', encoding='utf-8') as f:
    admin_data = json.load(f)

properties = admin_data.get('properties', [])

for p in properties:
    pid = str(p.get('id'))
    name = p.get('name', '')
    sdate = p.get('start_date', '')
    
    if pid == '24' or '203' in name:
        print(f"Updating APTO 203 (ID {pid}): setting JULIO status to NEW_CONTRACT")
        # Update 2026 and 2027 JULIO status
        for y in ['2026', '2027']:
            if y in p.get('payments', {}):
                for m in p['payments'][y]:
                    if m.get('month') == 'JULIO':
                        m['status'] = 'NEW_CONTRACT'
                        m['value'] = 'CONTRATO NUEVO'

# Save admin_data.json
with open('admin_data.json', 'w', encoding='utf-8') as f:
    json.dump(admin_data, f, indent=4, ensure_ascii=False)

print("Saved admin_data.json!")

# Update Base de datos Admin.xlsx
wb = openpyxl.load_workbook('Base de datos Admin.xlsx')
ws = wb['ADMINISTRACION DETALLADA']

# Row 25 is APTO 203
# Col for JULIO 2026 is col 63 (index 57 + 6)
# 2026 starts at col 57: ENE=57, FEB=58, MAR=59, ABR=60, MAY=61, JUN=62, JUL=63
for r in range(5, ws.max_row + 1):
    pname = str(ws.cell(r, 9).value or '')
    if '203' in pname:
        print(f"Updating Excel row {r} for APTO 203 JULIO 2026 & 2027 to CONTRATO NUEVO")
        ws.cell(r, 63, value='CONTRATO NUEVO') # 2026 JULIO
        ws.cell(r, 76, value='CONTRATO NUEVO') # 2027 JULIO

wb.save('Base de datos Admin.xlsx')
print("Saved Base de datos Admin.xlsx!")

# Send to Google Apps Script cloud
url = "https://script.google.com/macros/s/AKfycbwAUUSYRhDX6Eik4KA-B6luk74YjCNRanwv13CmmZg4La8NzVuNyBC0T5GH6f4-ke-Xig/exec"
payload = {
    "action": "importAdminData",
    "data": {
        "properties": properties
    }
}

try:
    res = requests.post(url, json=payload, timeout=20)
    print("Google Apps Script POST response:", res.status_code, res.text)
except Exception as e:
    print("Error posting to Google Apps Script:", e)
