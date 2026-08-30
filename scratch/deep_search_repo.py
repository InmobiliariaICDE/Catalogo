import os
import json
import openpyxl

print("=== SEARCHING ENTIRE REPOSITORY FOR PORTAL AND NOGAL ===")
root_dir = "."
matches = []

for root, dirs, files in os.walk(root_dir):
    if '.git' in root or 'node_modules' in root or 'scratch' in root or '.system_generated' in root:
        continue
    for file in files:
        filepath = os.path.join(root, file)
        ext = os.path.splitext(file)[1].lower()
        
        if ext in ['.json', '.txt', '.js', '.html', '.csv', '.py', '.gs', '.md']:
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    for kw in ['portal del campo', 'portal campo', 'nogales', 'nogal']:
                        if kw in content.lower():
                            matches.append((filepath, kw, content.lower().count(kw)))
            except Exception as e:
                pass
        elif ext in ['.xlsx']:
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for r in range(1, ws.max_row + 1):
                        row_str = " ".join([str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]).lower()
                        for kw in ['portal del campo', 'portal campo', 'nogales', 'nogal']:
                            if kw in row_str:
                                matches.append((f"{filepath} [{sheet} row {r}]", kw, 1))
            except Exception as e:
                pass

print(f"Total matches found: {len(matches)}")
for m in matches:
    print(f"File/Location: {m[0]} | Keyword: {m[1]} | Occurrences: {m[2]}")
