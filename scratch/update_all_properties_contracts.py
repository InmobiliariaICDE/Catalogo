import json
import openpyxl
import requests

with open('admin_data.json', 'r', encoding='utf-8') as f:
    admin_data = json.load(f)

properties = admin_data.get('properties', [])

for p in properties:
    pid = str(p.get('id'))
    if pid == '11': # APTO 101
        p['start_date'] = '2024-10-05'
        p['duration'] = '12'
        print("Updated ID 11 (APTO 101) start_date to 2024-10-05")
    elif pid == '24': # APTO 203
        p['tenant_name'] = 'Lucas Garcia Pedraza'
        p['tenant_phone'] = '3144049521'
        p['start_date'] = '2026-08-01'
        p['duration'] = '12'
        p['deposit'] = '300000'
        p['monthly_rent'] = 450000.0
        p['due_day'] = 26.0
        p['max_due_day'] = 30.0
        print("Updated ID 24 (APTO 203) with Lucas Garcia Pedraza, start_date 2026-08-01")

# Save admin_data.json
with open('admin_data.json', 'w', encoding='utf-8') as f:
    json.dump(admin_data, f, indent=4, ensure_ascii=False)

print("Saved admin_data.json!")

# Update Base de datos Admin.xlsx
wb = openpyxl.load_workbook('Base de datos Admin.xlsx')
ws = wb['ADMINISTRACION DETALLADA']

for r in range(6, ws.max_row + 1):
    pid = ws.cell(r, 1).value
    if str(pid) == '11':
        ws.cell(r, 14, value='2024-10-05')
    elif str(pid) == '24':
        ws.cell(r, 10, value='Lucas Garcia Pedraza')
        ws.cell(r, 11, value='3144049521')
        ws.cell(r, 12, value='12')
        ws.cell(r, 13, value='300000')
        ws.cell(r, 14, value='2026-08-01')
        ws.cell(r, 15, value=26)
        ws.cell(r, 16, value=30)
        ws.cell(r, 17, value=450000)

wb.save('Base de datos Admin.xlsx')
print("Saved Base de datos Admin.xlsx!")

# POST to Google Apps Script
url = "https://script.google.com/macros/s/AKfycbwAUUSYRhDX6Eik4KA-B6luk74YjCNRanwv13CmmZg4La8NzVuNyBC0T5GH6f4-ke-Xig/exec"
payload = {
    "action": "importAdminData",
    "data": {
        "properties": properties
    }
}

try:
    res = requests.post(url, json=payload, timeout=20)
    print("Google Apps Script response:", res.status_code, res.text)
except Exception as e:
    print("Error posting to Google Apps Script:", e)
