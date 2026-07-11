import openpyxl
import subprocess

file_path = "Base de datos Admin.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['ADMINISTRACION DETALLADA']

# Row 22 is index 21, column 63 is JULIO (2026)
cell_val = ws.cell(row=22, column=63).value
print(f"Original cell value: {cell_val}")

# Set the cell to simulate the user's July payment
ws.cell(row=22, column=63).value = 740000

wb.save(file_path)
print("Simulated July payment saved to Base de datos Admin.xlsx!")

# Run data update
subprocess.run(["python", "actualizar_admin.py"])
