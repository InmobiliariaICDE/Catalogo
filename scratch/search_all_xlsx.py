import openpyxl
import os
import glob

print("=== CHECKING ALL XLSX FILES FOR PORTAL, CAMPO, NOGAL, NOGALES ===")
xlsx_files = glob.glob("*.xlsx")

for fx in xlsx_files:
    print(f"\n--- Checking File: {fx} ---")
    try:
        wb = openpyxl.load_workbook(fx, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            match_count = 0
            for r in range(1, ws.max_row + 1):
                row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None]
                row_str = " ".join([str(v) for v in row_vals]).lower()
                for kw in ['portal del campo', 'portal', 'nogal', 'nogales', 'campo']:
                    if kw in row_str:
                        match_count += 1
                        print(f"  [{sname} Row {r}] ({kw}): {row_str[:160]}")
                        break
            if match_count > 0:
                print(f"  -> Total matches in sheet '{sname}': {match_count}")
    except Exception as e:
        print(f"Error reading {fx}:", e)
