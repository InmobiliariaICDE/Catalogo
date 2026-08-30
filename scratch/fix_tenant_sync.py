import json
import openpyxl

# Update admin_data.json
with open('admin_data.json', 'r', encoding='utf-8') as f:
    admin_data = json.load(f)

properties = admin_data.get('properties', [])

for p in properties:
    pid = str(p.get('id'))
    pname = p.get('name', '').upper()
    if 'LIMONAR' in pname:
        p['tenant_name'] = 'Elsa Oviedo Murcia'
        p['tenant_phone'] = '3229816445'
    elif 'NOGALES' in pname:
        p['tenant_name'] = '' # Desocupado
    elif 'PORTAL DEL CAMPO' in pname:
        p['tenant_name'] = '' # Desocupado

# Save admin_data.json
with open('admin_data.json', 'w', encoding='utf-8') as f:
    json.dump(admin_data, f, indent=4, ensure_ascii=False)

print("Updated admin_data.json!")

# Update Base de datos Admin.xlsx
wb = openpyxl.load_workbook('Base de datos Admin.xlsx')
ws = wb['ADMINISTRACION DETALLADA']

for r in range(6, ws.max_row + 1):
    pname = str(ws.cell(r, 9).value or '').upper()
    if 'LIMONAR' in pname:
        ws.cell(r, 10, value='Elsa Oviedo Murcia')
        ws.cell(r, 11, value='3229816445')

wb.save('Base de datos Admin.xlsx')
print("Updated Base de datos Admin.xlsx!")
