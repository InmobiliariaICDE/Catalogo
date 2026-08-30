import openpyxl

wb = openpyxl.load_workbook("Base de datos Admin.xlsx")
sheet = wb.active

print("Sheet max row:", sheet.max_row)
for r in range(15, sheet.max_row + 1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 18)]
    print(f"Row {r:2d}: {vals}")
