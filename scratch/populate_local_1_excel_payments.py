import openpyxl, json

with open("recovered_local_1.json", encoding="utf-8") as f:
    local_1 = json.load(f)

wb = openpyxl.load_workbook("Base de datos Admin.xlsx")
sheet = wb.active

# Find row for LOCAL 1
target_row = None
for r in range(6, sheet.max_row + 1):
    val = str(sheet.cell(row=r, column=9).value)
    if 'LOCAL 1' in val.upper():
        target_row = r
        break

if target_row:
    print(f"Populating payments for LOCAL 1 at row {target_row}...")
    year_cols = {
        '2023': 18,
        '2024': 31,
        '2025': 44,
        '2026': 57,
        '2027': 70
    }
    
    payments = local_1.get('payments', {})
    for yr, m_list in payments.items():
        if yr in year_cols:
            start_col = year_cols[yr]
            for m_idx, m_obj in enumerate(m_list):
                val = m_obj.get('value')
                col_idx = start_col + m_idx
                sheet.cell(row=target_row, column=col_idx, value=val)
                
    wb.save("Base de datos Admin.xlsx")
    print("SUCCESS: Payments for LOCAL 1 populated in Base de datos Admin.xlsx!")
else:
    print("ERROR: LOCAL 1 row not found in Excel sheet.")
