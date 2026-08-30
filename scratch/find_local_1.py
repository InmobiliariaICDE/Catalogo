import openpyxl, json

wb = openpyxl.load_workbook("Base de datos Admin.xlsx", data_only=True)
sheet = wb.active

print("Searching for LOCAL 1 in Base de datos Admin.xlsx:")
for r in range(1, sheet.max_row + 1):
    row_vals = [str(sheet.cell(row=r, column=c).value) for c in range(1, min(20, sheet.max_column + 1))]
    line = " | ".join(row_vals)
    if 'LOCAL 1' in line.upper():
        print(f"Row {r:2d}: {line}")

d = json.load(open('admin_data.json', encoding='utf-8'))
print(f"\nTotal properties in current admin_data.json: {len(d['properties'])}")
for p in d['properties']:
    if 'LOCAL' in p['name'].upper():
        print("Found in admin_data.json:", p['name'])
