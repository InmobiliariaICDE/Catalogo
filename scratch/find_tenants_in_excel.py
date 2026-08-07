import openpyxl

wb = openpyxl.load_workbook("Base de datos Admin.xlsx", data_only=True)
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"--- Sheet: {sheetname} ({ws.max_row} rows, {ws.max_column} cols) ---")
    for r in range(1, min(40, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=r, column=c).value) for c in range(1, min(25, ws.max_column + 1))]
        line = " | ".join(row_vals)
        if any(name in line.upper() for name in ['DIANA', 'NATALIA', 'VANAGAS', 'LIZETH', 'LUIS', 'GIL', 'DARWIN', 'VANEGAS', 'GUTIERREZ']):
            print(f"Row {r:2d}: {line}")
