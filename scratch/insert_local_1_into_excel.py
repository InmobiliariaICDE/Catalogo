import openpyxl

wb = openpyxl.load_workbook("Base de datos Admin.xlsx")
sheet = wb.active

# Check if LOCAL 1 is already in sheet
found = False
for r in range(1, 30):
    val = str(sheet.cell(row=r, column=9).value)
    if 'LOCAL 1' in val.upper():
        found = True
        print(f"LOCAL 1 already in sheet row {r}")
        break

if not found:
    sheet.insert_rows(21)
    sheet.cell(row=21, column=1, value=20)
    sheet.cell(row=21, column=2, value=16)
    sheet.cell(row=21, column=3, value=20)
    sheet.cell(row=21, column=4, value=16)
    sheet.cell(row=21, column=7, value="Silvia")
    sheet.cell(row=21, column=9, value="LOCAL 1 1 | Aumento 24 abril 2025 | $35.000 | 2. Aumento 2027")
    sheet.cell(row=21, column=12, value=6)
    sheet.cell(row=21, column=13, value="400000")
    sheet.cell(row=21, column=14, value="2023-04-24")
    sheet.cell(row=21, column=15, value=24)
    sheet.cell(row=21, column=16, value=29)
    sheet.cell(row=21, column=17, value=535000)
    
    wb.save("Base de datos Admin.xlsx")
    print("SUCCESS: LOCAL 1 inserted into Base de datos Admin.xlsx row 21!")
