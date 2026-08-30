import subprocess
import json
import openpyxl
import os

GIT = r"C:\Users\USUARIO\AppData\Local\GitHubDesktop\app-3.6.3\resources\app\git\cmd\git.exe"

def run_git(args):
    res = subprocess.run([GIT] + args, capture_output=True)
    return res.stdout.decode('utf-8', errors='ignore')

print("=== 1. ALL COMMITS IN REPO ===")
stdout_all = run_git(["log", "--oneline", "-n", "100"])
all_commits = [l.split()[0] for l in stdout_all.strip().splitlines() if l]
print(f"Total commits analyzed: {len(all_commits)}")

found_in_git = []
for c in all_commits:
    c_text = run_git(["show", c]).lower()
    for kw in ['portal del campo', 'portal campo', 'nogales', 'nogal', '21 inmuebles', '21 propiedades']:
        if kw in c_text:
            found_in_git.append((c, kw))

print(f"Git diff keyword matches: {len(found_in_git)}")
for c, kw in found_in_git:
    print(f"  Commit {c} matches '{kw}'")

print("\n=== 2. EXCEL FULL ROW INSPECTION IN Base de datos Admin.xlsx ===")
wb = openpyxl.load_workbook('Base de datos Admin.xlsx', data_only=True)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"Sheet '{sname}' max_row={ws.max_row}, max_col={ws.max_column}")
    non_empty_rows = 0
    for r in range(1, ws.max_row + 1):
        r_vals = [ws.cell(r, col).value for col in range(1, 20)]
        if any(v is not None for v in r_vals):
            non_empty_rows += 1
            r_str = " | ".join([str(v) for v in r_vals if v is not None])
            if any(k in r_str.lower() for k in ['portal', 'nogal', 'campo', '21']):
                print(f"  Row {r} match in {sname}: {r_str[:150]}")
    print(f"Sheet '{sname}' non-empty rows count: {non_empty_rows}")

print("\n=== 3. CHECK ALL HISTORICAL admin_data.json VERSIONS ===")
stdout_admin_commits = run_git(["log", "--oneline", "--", "admin_data.json"])
admin_commits = [l.split()[0] for l in stdout_admin_commits.strip().splitlines() if l]

all_historical_prop_names = set()
for c in admin_commits:
    json_str = run_git(["show", f"{c}:admin_data.json"])
    if json_str:
        try:
            d = json.loads(json_str)
            props = d.get("properties", [])
            print(f"Commit {c}: count={len(props)}")
            for p in props:
                name = p.get("name", "")
                owner = p.get("owner", "")
                all_historical_prop_names.add(f"ID={p.get('id')} | Name='{name}' | Owner='{owner}'")
        except Exception:
            pass

print("\nUnique historical properties in admin_data.json git history:")
for hp in sorted(all_historical_prop_names):
    print(" ", hp)
