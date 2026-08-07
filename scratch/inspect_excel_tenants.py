import openpyxl

wb = openpyxl.load_workbook("Base de datos Admin.xlsx", data_only=True)
sheet = wb.active

print("Headers row 5:")
for col in range(1, 20):
    val = sheet.cell(row=5, column=col).value
    print(f"Col {col} ({openpyxl.utils.get_column_letter(col)}): '{val}'")

print("\nSample rows (6 to 15):")
for r in range(6, 16):
    col_a = sheet.cell(row=r, column=1).value
    col_i = sheet.cell(row=r, column=9).value # Property name
    col_j = sheet.cell(row=r, column=10).value # Tenant?
    col_k = sheet.cell(row=r, column=11).value # Tenant phone?
    print(f"Row {r:2d} | Col A: {str(col_a):<5} | Col I (Name): {str(col_i):<35} | Col J (Tenant): {str(col_j):<30} | Col K (Phone): {str(col_k)}")
