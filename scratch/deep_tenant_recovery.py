import os
import json
import openpyxl
import re
import subprocess

GIT = r"C:\Users\USUARIO\AppData\Local\GitHubDesktop\app-3.6.3\resources\app\git\cmd\git.exe"

def run_git(args):
    res = subprocess.run([GIT] + args, capture_output=True)
    return res.stdout.decode('utf-8', errors='ignore')

print("=== 1. SEARCHING ALL EXCEL FILES FOR NAMES ===")
for root, dirs, files in os.walk('.'):
    if '.git' in root or 'node_modules' in root or 'scratch' in root:
        continue
    for file in files:
        if file.endswith('.xlsx'):
            path = os.path.join(root, file)
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for r in range(1, ws.max_row + 1):
                        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None]
                        r_str = " | ".join([str(v) for v in vals])
                        if any(p in r_str.lower() for p in ['apto', 'casa', 'local', 'limonar', 'goya', 'lilola', 'manzanares', 'chapinero', 'marcos', 'silvia', 'nogales', 'portal']):
                            # Look for names in this row
                            for word in r_str.split('|'):
                                w = word.strip()
                                if len(w) > 3 and not any(char.isdigit() for char in w) and not '$' in w:
                                    if any(title_word in w for title_word in ['Maria', 'Jose', 'Carlos', 'Ana', 'Luis', 'Juan', 'Diego', 'Sandra', 'Patricia', 'Nini', 'Angela', 'Marcos', 'Jorge', 'Eduard', 'Giovany', 'Nohora', 'Silvia', 'Elsa', 'Joel', 'Vargas', 'Motta', 'Renza', 'Dussan', 'Hurtado']):
                                        print(f"File: {path} [{sheet} R{r}]: {w} <FULL ROW: {r_str[:120]}>")
            except Exception as e:
                pass

print("\n=== 2. SEARCHING ALL GIT COMMITS FOR PROPERTY-NAME PAIRS ===")
commits_raw = run_git(["log", "-n", "100", "--oneline"])
commits = [l.split()[0] for l in commits_raw.strip().splitlines() if l]

for c in commits:
    diff = run_git(["show", c])
    for kw in ['tenant', 'inquilino', 'arrendatario', 'arriendo']:
        if kw in diff.lower():
            for line in diff.splitlines():
                if any(k in line.lower() for k in ['tenant_name', 'inquilino', 'arrendatario']) and ('+' in line or '-' in line):
                    print(f"Commit {c} line: {line.strip()[:140]}")
