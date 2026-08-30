import json
import openpyxl

# Read admin_data.json
with open('admin_data.json', 'r', encoding='utf-8') as f:
    admin_data = json.load(f)

for p in admin_data.get('properties', []):
    if '203' in p.get('name', ''):
        print("JSON property:", json.dumps(p, indent=2, ensure_ascii=False))

# Read Base de datos Admin.xlsx
wb = openpyxl.load_workbook('Base de datos Admin.xlsx', data_only=True)
ws = wb['ADMINISTRACION DETALLADA']

for r in range(5, ws.max_row + 1):
    pname = str(ws.cell(r, 9).value or '')
    if '203' in pname:
        row_vals = [ws.cell(r, c).value for c in range(1, 18)]
        print(f"Excel Row {r}: {row_vals}")
