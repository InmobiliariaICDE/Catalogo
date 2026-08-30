import json
import openpyxl

with open('admin_data.json', 'r', encoding='utf-8') as f:
    d = json.load(f)

props = d.get('properties', [])
print(f"admin_data.json count: {len(props)} properties")
for i, p in enumerate(props):
    print(f" {i+1:2d}. ID {p.get('id'):<4} | Name: {p.get('name'):<38} | Owner: {p.get('owner')}")

wb = openpyxl.load_workbook('Base de datos Admin.xlsx', data_only=True)
ws = wb['ADMINISTRACION DETALLADA']

excel_count = 0
for r in range(6, ws.max_row + 1):
    if ws.cell(r, 1).value is not None and ws.cell(r, 9).value is not None:
        excel_count += 1

print(f"\nBase de datos Admin.xlsx count: {excel_count} properties")
