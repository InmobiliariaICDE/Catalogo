import json
import openpyxl
import re
import os

print("=== 1. TENANT NAMES IN dashboard_inmobiliario.html ===")
with open('ADMINISTRACION/dashboard_inmobiliario.html', 'r', encoding='utf-8', errors='ignore') as f:
    dash_html = f.read()

# Extract table rows
rows = re.findall(r'<tr>(.*?)</tr>', dash_html, re.DOTALL)
for r in rows:
    clean = re.sub(r'<.*?>', ' | ', r)
    clean = " ".join(clean.split())
    if any(k in clean.lower() for k in ['apto', 'casa', 'local', 'limonar', 'lilola', 'goya']):
        print("Dash row:", clean)

print("\n=== 2. TENANT NAMES IN INGRESO Y GASTOS FIJOS .xlsx ===")
if os.path.exists('INGRESO Y GASTOS FIJOS .xlsx'):
    wb = openpyxl.load_workbook('INGRESO Y GASTOS FIJOS .xlsx', data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, min(40, ws.max_row + 1)):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None]
            if row_vals:
                row_str = " | ".join([str(v) for v in row_vals])
                if any(k in row_str.lower() for k in ['arriend', 'inquilino', 'arrendatario', 'pag', 'canon']):
                    print(f"[{sname} R{r}]: {row_str[:160]}")
