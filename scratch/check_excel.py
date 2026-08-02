import openpyxl, glob

for fpath in glob.glob('*.xlsx'):
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for r in sheet.iter_rows(values_only=True):
                for cell in r:
                    if cell is not None and ('883' in str(cell) or '883750' in str(cell)):
                        print(f"Found in {fpath} sheet {sname}: {cell}")
    except Exception as e:
        print(f"Error reading {fpath}: {e}")
