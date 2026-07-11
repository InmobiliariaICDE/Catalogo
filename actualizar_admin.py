import pandas as pd
import json
import os
import re
import sys
import requests
import openpyxl
from datetime import datetime

def clean_prop_name(raw_name):
    if not raw_name or pd.isna(raw_name):
        return "", ""
    raw_name = str(raw_name).strip()
    
    # Split by two or more spaces, or by a common marker like "1. Aumento" or "Aumento"
    parts = re.split(r'\s{2,}|(?=1\.\s*Aumento)|(?=Aumento)', raw_name)
    name = parts[0].strip()
    
    # Remove any trailing dots
    if name.endswith('.'):
        name = name[:-1].strip()
        
    notes = " | ".join([p.strip() for p in parts[1:] if p.strip()])
    return name, notes

def parse_date(val):
    if pd.isna(val):
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    if "00:00:00" in val_str:
        val_str = val_str.split(" ")[0]
    return val_str

def parse_number(val):
    if pd.isna(val) or val == "" or val == "-":
        return 0
    try:
        # Remove currency symbols and thousand separators
        cleaned = str(val).replace('$', '').replace('.', '').replace(',', '').strip()
        return float(cleaned)
    except:
        return 0

def get_month_status(val, year, month_idx, start_date_str, due_day, monthly_rent):
    # Current system date for comparison
    today = datetime.now()
    current_year = today.year
    current_month_idx = today.month - 1  # 0-indexed: 0 = Enero, 1 = Febrero...
    
    if pd.isna(val):
        val_str = "-"
    else:
        val_str = str(val).strip().upper()
        
    # Check if vacant/preaviso/etc. based on string values
    if "DESOCUPAD" in val_str:
        return "VACANT", val_str
    elif "PREAVISO" in val_str:
        return "PREAVISO", val_str
    elif "NUEVO" in val_str or "CONTRATO NUEVO" in val_str:
        return "NEW_CONTRACT", val_str
    elif "NO RENOVARA" in val_str:
        return "NO_RENEW", val_str
    elif "ENTREGA" in val_str:
        return "DELIVERY", val_str
    
    # If the value is a positive number
    num_val = parse_number(val)
    if num_val > 0:
        return "PAID", num_val
        
    # If the value is 0 or empty/dash
    # 1. Check if the month is in the future
    if year > current_year or (year == current_year and month_idx > current_month_idx):
        return "FUTURE", val_str
        
    # If it is the current month but before the due day, treat as AL_DIA (Al día)
    if year == current_year and month_idx == current_month_idx:
        today_day = today.day
        limit_day = due_day if (due_day and due_day > 0) else 1
        if today_day < limit_day:
            return "AL_DIA", val_str
        
    # 2. Check if the contract had not started yet
    if start_date_str:
        try:
            # Parse start date (formats: YYYY-MM-DD or D-month-YY)
            # Simplistic check: if start date year/month is after this year/month
            start_dt = None
            for fmt in ("%Y-%m-%d", "%d-%b-%y"):
                try:
                    start_dt = datetime.strptime(start_date_str, fmt)
                    break
                except:
                    pass
            if start_dt:
                # Compare year and month
                if start_dt.year > year or (start_dt.year == year and start_dt.month > (month_idx + 1)):
                    return "UNSTARTED", val_str
        except Exception as e:
            pass
            
    # Default is pending
    return "PENDING", val_str

def parse_properties(file):
    if not os.path.exists(file):
        print(f"File {file} not found")
        return []
        
    df = pd.read_excel(file, sheet_name='ADMINISTRACION DETALLADA', header=None)
    properties = []
    
    # Loop from row index 5 through the end of the sheet dynamically
    for i in range(5, df.shape[0]):
        row = df.iloc[i]
        if len(row) < 15:
            continue
            
        # Parse basic fields according to Base de datos Admin.xlsx columns
        row_id = str(row[0]).strip() if not pd.isna(row[0]) else str(i-4)
        damage_notes = str(row[5]).strip() if not pd.isna(row[5]) else ""
        owner = str(row[6]).strip() if not pd.isna(row[6]) else "Sin Propietario"
        owner_phone = str(row[7]).strip() if not pd.isna(row[7]) else ""
        raw_name = str(row[8]).strip() if not pd.isna(row[8]) else ""
        
        if not raw_name or raw_name.lower() == 'nan':
            continue
            
        prop_name, increase_notes = clean_prop_name(raw_name)
        
        tenant_name = str(row[9]).strip() if not pd.isna(row[9]) else ""
        tenant_phone = str(row[10]).strip() if not pd.isna(row[10]) else ""
        
        duration = str(row[11]).strip() if not pd.isna(row[11]) else ""
        deposit = str(row[12]).strip() if not pd.isna(row[12]) else ""
        start_date = parse_date(row[13])
        due_day = parse_number(row[14])
        max_due_day = parse_number(row[15])
        monthly_rent = parse_number(row[16])
        
        # Payment columns mappings
        years_map = {
            2023: list(range(17, 29)),
            2024: list(range(30, 42)),
            2025: list(range(43, 55)),
            2026: list(range(56, 68)),
            2027: list(range(69, 81))
        }
        
        months_names = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        
        payments_history = {}
        for year, cols in years_map.items():
            payments_history[str(year)] = []
            for m_idx, col_idx in enumerate(cols):
                cell_val = row[col_idx]
                status, clean_val = get_month_status(cell_val, year, m_idx, start_date, due_day, monthly_rent)
                payments_history[str(year)].append({
                    "month": months_names[m_idx],
                    "value": clean_val,
                    "status": status
                })
        
        # Overall status is usually Occupied, unless latest current month status is VACANT
        overall_status = "Ocupado"
        if "DESOCUPAD" in raw_name.upper():
            overall_status = "Desocupado"
        else:
            # Check current month payment status
            from datetime import datetime
            _today = datetime.now()
            _curr_year_str = str(_today.year)
            _curr_month_idx = _today.month - 1
            if _curr_year_str in payments_history and len(payments_history[_curr_year_str]) > _curr_month_idx:
                _curr_pay = payments_history[_curr_year_str][_curr_month_idx]
                if _curr_pay["status"] == "VACANT":
                    overall_status = "Desocupado"

        if overall_status == "Desocupado":
            for year_str in payments_history:
                for m in payments_history[year_str]:
                    if m["status"] in ("PENDING", "AL_DIA", "FUTURE") and m["value"] in ("-", ""):
                        m["status"] = "VACANT"
                        m["value"] = "DESOCUPADO"
                        
        if overall_status == "Ocupado":
            today = datetime.now()
            current_year = today.year
            current_month_idx = today.month - 1
            for year_str in payments_history:
                for m_idx, m in enumerate(payments_history[year_str]):
                    y = int(year_str)
                    is_current = (y == current_year and m_idx == current_month_idx)
                    is_future = (y > current_year or (y == current_year and m_idx > current_month_idx))
                    
                    if m["status"] == "VACANT" and (is_current or is_future):
                        if is_current:
                            today_day = today.day
                            limit_day = due_day if (due_day and due_day > 0) else 1
                            if today_day < limit_day:
                                m["status"] = "AL_DIA"
                            else:
                                m["status"] = "PENDING"
                        else:
                            m["status"] = "FUTURE"
        # Propagar desocupado después de una entrega
        is_vacant_after_delivery = False
        sorted_years = sorted(payments_history.keys(), key=int)
        for year_str in sorted_years:
            for m in payments_history[year_str]:
                if m["status"] == "DELIVERY":
                    is_vacant_after_delivery = True
                elif m["status"] in ("PAID", "NEW_CONTRACT"):
                    is_vacant_after_delivery = False
                elif is_vacant_after_delivery:
                    if m["status"] in ("PENDING", "AL_DIA", "FUTURE", "UNSTARTED"):
                        m["status"] = "VACANT"
                        m["value"] = "DESOCUPADO"

        properties.append({
            "id": row_id,
            "excel_row": i,
            "owner": owner,
            "owner_phone": owner_phone,
            "name": prop_name,
            "increase_notes": increase_notes,
            "tenant_name": tenant_name,
            "tenant_phone": tenant_phone,
            "damage_notes": damage_notes,
            "duration": duration,
            "deposit": deposit,
            "start_date": start_date,
            "due_day": due_day,
            "max_due_day": max_due_day,
            "monthly_rent": monthly_rent,
            "status": overall_status,
            "payments": payments_history
        })
        
    return properties

def parse_silvia_ledger(file):
    if not os.path.exists(file):
        print(f"File {file} not found")
        return {}
        
    xl = pd.ExcelFile(file)
    silvia_data = {}
    
    for sheet in xl.sheet_names:
        if "ADMT - SILVIA" in sheet:
            year = sheet.split(" ")[-1]
            df = pd.read_excel(file, sheet_name=sheet, header=None)
            
            # The transaction table usually starts at row index 6 (row index 5 has headers)
            # Find the header row
            header_row = 5
            transactions = []
            
            for r_idx in range(header_row + 1, df.shape[0]):
                row = df.iloc[r_idx]
                
                # If Col4 (Mes/Descripción) is completely empty, skip
                desc = str(row[4]).strip() if not pd.isna(row[4]) else ""
                if not desc or desc == "nan":
                    continue
                    
                date = parse_date(row[3])
                recaudo = parse_number(row[5])
                abono = parse_number(row[6])
                saldo = parse_number(row[7])
                
                transactions.append({
                    "date": date,
                    "description": desc,
                    "recaudo": recaudo,
                    "abono": abono,
                    "saldo": saldo
                })
                
            silvia_data[year] = transactions
            
    return silvia_data

def pull_from_cloud():
    print("Obteniendo datos actualizados desde la nube...")
    url = "https://script.google.com/macros/s/AKfycbwAUUSYRhDX6Eik4KA-B6luk74YjCNRanwv13CmmZg4La8NzVuNyBC0T5GH6f4-ke-Xig/exec?action=getAdminData"
    try:
        res = requests.get(url, timeout=15)
        if res.status_code != 200:
            print(f"Error al conectar con la nube: Código HTTP {res.status_code}")
            return False
        
        cloud_data = res.json()
        if "error" in cloud_data:
            print("Error devuelto por Apps Script:", cloud_data["error"])
            return False
            
        properties = cloud_data.get("properties", [])
        if not properties:
            print("No se encontraron propiedades en los datos de la nube.")
            return False
            
        print(f"Se obtuvieron {len(properties)} propiedades de la nube. Actualizando Excel local en el lugar...")
        
        file_path = "Base de datos Admin.xlsx"
        if not os.path.exists(file_path):
            print(f"Error: El archivo {file_path} no existe.")
            return False
            
        # Abrir el Excel con openpyxl preservando fórmulas y estilos
        wb = openpyxl.load_workbook(file_path, data_only=False)
        sheet_name = 'ADMINISTRACION DETALLADA'
        if sheet_name not in wb.sheetnames:
            found = False
            for name in wb.sheetnames:
                if 'ADMINISTRACION' in name.upper():
                    sheet_name = name
                    found = True
                    break
            if not found:
                print(f"No se encontró la hoja '{sheet_name}' en el Excel.")
                return False
                
        ws = wb[sheet_name]
        
        # Mapeo de columnas por año (1-indexed para openpyxl)
        years_map = {
            2023: 18,
            2024: 31,
            2025: 44,
            2026: 57,
            2027: 70
        }
        
        updated_cells_count = 0
        
        for prop in properties:
            prop_id = prop.get("id")
            prop_name = prop.get("name", "")
            payments_history = prop.get("payments", {})
            
            # Buscar la fila correspondiente
            row_idx = -1
            for r in range(6, ws.max_row + 1):
                cell_id = ws.cell(row=r, column=1).value
                if cell_id is not None and str(cell_id).strip() == str(prop_id).strip():
                    row_idx = r
                    break
                    
            if row_idx == -1 and prop_name:
                # Buscar por coincidencia aproximada de nombre
                clean_prop_name = str(prop_name).strip().lower()
                for r in range(6, ws.max_row + 1):
                    cell_name = ws.cell(row=r, column=7).value
                    if cell_name is not None:
                        clean_cell_name = str(cell_name).strip().lower()
                        if clean_prop_name in clean_cell_name or clean_cell_name in clean_prop_name:
                            row_idx = r
                            break
                            
            if row_idx == -1:
                print(f"No se pudo encontrar la fila para el inmueble: ID={prop_id}, Nombre={prop_name}")
                continue
                
            # Actualizar pagos de esta propiedad
            for year_str, months_list in payments_history.items():
                year = int(year_str)
                if year not in years_map:
                    continue
                start_col = years_map[year]
                
                for month_idx, month_pay in enumerate(months_list):
                    val = month_pay.get("value", "-")
                    col_idx = start_col + month_idx
                    
                    # Leer valor actual de la celda en Excel
                    current_cell = ws.cell(row=row_idx, column=col_idx)
                    current_val = current_cell.value
                    
                    # Normalizar valores para comparación
                    excel_val_str = "" if current_val is None else str(current_val).strip().upper()
                    cloud_val_str = "" if val is None or val == "-" else str(val).strip().upper()
                    
                    # Comparación inteligente (número vs texto)
                    try:
                        cloud_num = float(val)
                        excel_num = 0.0
                        if current_val is not None:
                            try:
                                cleaned_excel = str(current_val).replace('$', '').replace('.', '').replace(',', '').strip()
                                excel_num = float(cleaned_excel)
                            except:
                                excel_num = 0.0
                        is_diff = abs(cloud_num - excel_num) > 0.01
                        final_write_val = cloud_num
                    except:
                        is_diff = excel_val_str != cloud_val_str
                        if val == "" or val == "-":
                            final_write_val = "-"
                        else:
                            final_write_val = val
                            
                    if is_diff:
                        current_cell.value = final_write_val
                        updated_cells_count += 1
                        print(f"Fila {row_idx}, Columna {col_idx} ({year_str} - Mes Index {month_idx}): Cambió '{current_val}' -> '{final_write_val}'")
                        
        if updated_cells_count > 0:
            wb.save(file_path)
            print(f"Éxito: Se actualizó el Excel local con {updated_cells_count} cambios de pagos.")
        else:
            print("El Excel local ya está sincronizado con la nube. No se requirieron cambios.")
        return True
    except Exception as e:
        print("Excepción al intentar sincronizar con la nube:", str(e))
        import traceback
        traceback.print_exc()
        return False

def main():
    print("Iniciando extracción y normalización de datos de administración...")
    
    # Soporte para jalar datos desde la nube (--pull)
    if "--pull" in sys.argv:
        pull_from_cloud()
        
    all_data = {
        "last_update": datetime.now().isoformat(),
        "properties": [],
        "silvia_ledger": {}
    }
    
    # 1. Parse Properties from Base de datos Admin.xlsx
    all_data["properties"] = parse_properties("Base de datos Admin.xlsx")
    
    # 2. Parse Silvia Ledger from Edif. Silvia - Pagos Admt. CRA7 No. 33-20.xlsx
    all_data["silvia_ledger"] = parse_silvia_ledger("Edif. Silvia - Pagos Admt. CRA7 No. 33-20.xlsx")
    
    # 3. Save to admin_data.json
    with open("admin_data.json", "w", encoding="utf-8") as f:
        json.dump(all_data, f, indent=4, ensure_ascii=False)
        
    print("Exito! Datos guardados en admin_data.json.")
    print(f"Total propiedades: {len(all_data['properties'])}")
    print(f"Anos en libro Silvia: {list(all_data['silvia_ledger'].keys())}")

if __name__ == "__main__":
    main()
